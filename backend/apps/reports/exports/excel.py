"""Excel export via openpyxl in write-only mode (no in-memory buffering
of the workbook rows). Money cells get number_format 'Rs. #,##0.00'
which lets users sum/filter natively."""

from __future__ import annotations

import io
from decimal import Decimal

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..base import Column, ReportResult


_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _to_cell_value(value, kind: str):
    if value is None:
        return ""
    if kind in ("money", "decimal", "int", "percent"):
        return Decimal(str(value)) if not isinstance(value, Decimal) else value
    return str(value) if kind == "text" else value


def _number_format(kind: str) -> str | None:
    return {
        "money": '"Rs."#,##0.00',
        "decimal": "#,##0.0000",
        "int": "#,##0",
        "percent": "0.00%",
    }.get(kind)


def excel_response(result: ReportResult, *, filename: str, sheet_name: str = "Report") -> HttpResponse:
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(sheet_name[:30])

    # Header row
    header = []
    for col in result.columns:
        from openpyxl.cell import WriteOnlyCell
        cell = WriteOnlyCell(ws, value=col.label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        header.append(cell)
    ws.append(header)

    # Data rows
    for row in result.rows:
        out = []
        for col in result.columns:
            from openpyxl.cell import WriteOnlyCell
            cell = WriteOnlyCell(ws, value=_to_cell_value(row.get(col.key), col.kind))
            fmt = _number_format(col.kind)
            if fmt is not None:
                cell.number_format = fmt
            out.append(cell)
        ws.append(out)

    # Set reasonable column widths
    for idx, col in enumerate(result.columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = max(12, len(col.label) + 4)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
