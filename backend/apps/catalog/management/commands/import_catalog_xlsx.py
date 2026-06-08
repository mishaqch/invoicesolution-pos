"""Bulk-import a product catalog from an .xlsx spreadsheet into a POS tenant.

Built for the "Pakistan POS Seed Catalog" sheet (cols: SKU, Category,
SubCategory, Brand, Product_Name, Barcode, PCT_HS_Code, UOM, Tax_Category) but
written defensively so any similarly-shaped sheet works.

What it does (and deliberately does NOT do):
  - Maps each row's (sub)category to a REAL PRAL HS code via SUBCATEGORY_HS
    (validated against the synced /pdi/v1/itemdesccode catalog). The sheet's own
    PCT_HS_Code is IGNORED when it's one of the known-bad placeholders (the seed
    file ships 1 HS code per category, several of which are wrong / FBR-rejected);
    we trust the subcategory map instead. If a row's sheet HS code IS a real PRAL
    code AND we have no subcategory mapping, we fall back to the sheet value.
  - DROPS the sheet's barcodes. The seed file's barcodes are sequential/synthetic
    (8964000000000, ...001) — not real GS1 codes — so scanning a physical product
    would never match. Barcodes are bound later via scan-to-assign. (--keep-barcodes
    overrides this if you ever import a sheet with genuine barcodes.)
  - Infers UOM from the product name (… 1L/500ml → LTR, loose/per kg → KG, else
    PCS). Falls back to the sheet UOM, then PCS.
  - Idempotent: get_or_create on (tenant, sku). Re-running adds only new rows.
  - POS tenants only (refuses digital_invoicing-only unless --force).

Usage:
    python manage.py import_catalog_xlsx --ntn 7886736-0 \
        --file /path/Pakistan_POS_Seed_Catalog_6000_Items.xlsx
    # options: --force  --keep-barcodes  --no-stock  --limit N (test a subset)
"""

from __future__ import annotations

import re
from decimal import Decimal

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.catalog.models import Category, HsCode, Product, TaxRate, UnitOfMeasure
from apps.inventory.services.movements import record_movement
from apps.tenants.models import Branch, Tenant

# Subcategory -> REAL PRAL HS code (every value verified to exist in PRAL).
# This is the source of truth for the HS code, NOT the spreadsheet column.
SUBCATEGORY_HS: dict[str, str] = {
    # Baby Care
    "Diapers": "9619.0010", "Formula": "1901.1000", "Wipes": "3401.1100",
    # Beverages
    "Coffee": "2101.1110", "Energy Drinks": "2202.9900", "Juices": "2009.8900",
    "Soft Drinks": "2202.1010", "Tea": "0902.4090", "Water": "2201.1010",
    # Dairy
    "Butter": "0405.1000", "Cheese": "0406.1010", "Milk": "0401.2000",
    "Yogurt": "0403.2000",
    # Electronics
    "Battery": "8506.5000", "Charger": "8504.4090", "Earphones": "8518.3000",
    "Power Bank": "8507.6000",
    # Frozen
    "Chicken": "0207.1400", "Fries": "2004.1000", "Nuggets": "1602.3200",
    "Paratha": "1905.9000",
    # Grocery
    "Flour": "1101.0010", "Pulses": "0713.6000", "Rice": "1006.3090",
    "Salt": "2501.0010", "Spices": "0910.9990", "Sugar": "1701.9920",
    # Household
    "Cleaner": "3402.9000", "Detergent": "3402.5000", "Dishwash": "3402.5000",
    "Tissue": "4818.1000",
    # Personal Care
    "Lotion": "3304.9990", "Shampoo": "3305.1000", "Soap": "3401.1100",
    "Toothpaste": "3306.1010",
    # Snacks
    "Biscuits": "1905.3100", "Chips": "2005.2000", "Crackers": "1905.3100",
    "Nimco": "2106.9090",
    # Stationery
    "Marker": "9608.2000", "Notebook": "4820.2000", "Pen": "9608.1000",
    "Pencil": "9609.1000",
}

# Default sale prices per subcategory (PKR) — the sheet has no price; give a
# sane placeholder the tenant edits. Cost defaults to ~80% of sale.
SUBCATEGORY_PRICE: dict[str, str] = {
    "Diapers": "45", "Formula": "1200", "Wipes": "300", "Coffee": "550",
    "Energy Drinks": "200", "Juices": "340", "Soft Drinks": "180", "Tea": "560",
    "Water": "90", "Butter": "480", "Cheese": "650", "Milk": "290",
    "Yogurt": "200", "Battery": "150", "Charger": "900", "Earphones": "800",
    "Power Bank": "2500", "Chicken": "650", "Fries": "450", "Nuggets": "550",
    "Paratha": "350", "Flour": "1100", "Pulses": "350", "Rice": "350",
    "Salt": "60", "Spices": "220", "Sugar": "150", "Cleaner": "350",
    "Detergent": "650", "Dishwash": "200", "Tissue": "200", "Lotion": "450",
    "Shampoo": "550", "Soap": "120", "Toothpaste": "260", "Biscuits": "120",
    "Chips": "120", "Crackers": "100", "Nimco": "150", "Marker": "120",
    "Notebook": "150", "Pen": "50", "Pencil": "30",
}

_LTR_RE = re.compile(r"(\d+(?:\.\d+)?)\s*(ml|l|ltr|litre|liter)\b", re.I)


def infer_uom(name: str, sheet_uom: str | None, uom_codes: set[str]) -> str:
    """Best-effort UoM from the product name; fall back to sheet, then PCS.

    Retail packs are sold as PCS even when labelled in ml/g (you sell ONE
    bottle, not 0.5 L), so we only pick LTR/KG for clearly loose/bulk items.
    """
    n = name.lower()
    if any(w in n for w in ("per kg", "loose", "/kg")) and "KG" in uom_codes:
        return "KG"
    if any(w in n for w in ("per litre", "per liter", "/l ")) and "LTR" in uom_codes:
        return "LTR"
    if sheet_uom and sheet_uom.upper() in uom_codes:
        return sheet_uom.upper()
    return "PCS"


class Command(BaseCommand):
    help = "Bulk-import a product catalog from an .xlsx into a POS tenant (real HS codes, no fake barcodes)."

    def add_arguments(self, parser):
        parser.add_argument("--ntn")
        parser.add_argument("--tenant-id")
        parser.add_argument("--file", required=True, help="Path to the .xlsx.")
        parser.add_argument("--force", action="store_true")
        parser.add_argument("--keep-barcodes", action="store_true",
                            help="Import the sheet's barcodes (only if they're REAL).")
        parser.add_argument("--no-stock", action="store_true")
        parser.add_argument("--limit", type=int, default=0,
                            help="Import only the first N rows (testing).")

    def handle(self, *args, **opts):
        try:
            import openpyxl
        except ImportError:
            raise CommandError("openpyxl not installed. pip install openpyxl")

        tenant = self._resolve_tenant(opts)
        if tenant.business_mode == "digital_invoicing" and not opts["force"]:
            raise CommandError(
                f"{tenant.business_name} is digital_invoicing-only. Use --force to override.")

        wb = openpyxl.load_workbook(opts["file"], read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = [str(h).strip() if h is not None else "" for h in next(rows)]
        col = {name: i for i, name in enumerate(header)}
        required = ["Category", "SubCategory", "Product_Name"]
        missing = [c for c in required if c not in col]
        if missing:
            raise CommandError(f"Sheet missing columns: {missing}. Found: {header}")

        uom_codes = set(UnitOfMeasure.objects.values_list("code", flat=True))
        # Validate every HS code we'll use exists, up-front.
        bad = sorted({c for c in SUBCATEGORY_HS.values()
                      if not HsCode.objects.filter(code=c).exists()})
        if bad:
            raise CommandError(
                "These mapped HS codes are missing — run sync_pral_reference:\n  "
                + "\n  ".join(bad))

        tax_rate, _ = TaxRate.objects.get_or_create(
            tenant=tenant, name="GST 18%",
            defaults={"rate": Decimal("18.00"), "is_default": True, "applies_to": "all"})
        branch = (Branch.objects.filter(tenant=tenant, deleted_at__isnull=True)
                  .order_by("-is_default", "name").first())

        cat_cache: dict[tuple[str, str | None], Category] = {}
        created = skipped = stocked = unmapped = 0
        n = 0
        # Batch in one transaction; 6k rows is fine.
        with transaction.atomic():
            for r in rows:
                if opts["limit"] and n >= opts["limit"]:
                    break
                n += 1
                cat = (r[col["Category"]] or "").strip()
                sub = (r[col["SubCategory"]] or "").strip()
                name = (r[col["Product_Name"]] or "").strip()
                if not name:
                    continue
                sku = str(r[col["SKU"]]).strip() if "SKU" in col and r[col["SKU"]] is not None else ""
                if not sku:
                    sku = re.sub(r"[^A-Za-z0-9]+", "-", name).upper()[:40]
                brand = (r[col["Brand"]] or "").strip() if "Brand" in col else ""

                # HS code: trust the subcategory map; else the sheet value IF real.
                hs = SUBCATEGORY_HS.get(sub)
                if not hs:
                    sheet_hs = str(r[col["PCT_HS_Code"]]).strip() if "PCT_HS_Code" in col and r[col["PCT_HS_Code"]] else ""
                    if sheet_hs and HsCode.objects.filter(code=sheet_hs).exists():
                        hs = sheet_hs
                    else:
                        unmapped += 1
                        continue  # never seed a product with an invalid HS code

                uom = infer_uom(name, r[col["UOM"]] if "UOM" in col else None, uom_codes)
                price = SUBCATEGORY_PRICE.get(sub, "100")
                cost = str((Decimal(price) * Decimal("0.8")).quantize(Decimal("1")))

                barcode = None
                if opts["keep_barcodes"] and "Barcode" in col and r[col["Barcode"]]:
                    barcode = str(r[col["Barcode"]]).strip()

                # Category (+ subcategory as a child category for navigation).
                parent = cat_cache.get((cat, None))
                if parent is None and cat:
                    parent, _ = Category.objects.get_or_create(
                        tenant=tenant, slug=_slug(cat), defaults={"name": cat})
                    cat_cache[(cat, None)] = parent
                category = cat_cache.get((cat, sub))
                if category is None and sub:
                    category, _ = Category.objects.get_or_create(
                        tenant=tenant, slug=_slug(f"{cat}-{sub}"),
                        defaults={"name": sub, "parent": parent})
                    cat_cache[(cat, sub)] = category
                category = category or parent

                _, was_created = Product.objects.get_or_create(
                    tenant=tenant, sku=sku,
                    defaults={
                        "name": name,
                        "name_ur": "",
                        "description": brand,
                        "category": category,
                        "uom": UnitOfMeasure.objects.get(code=uom),
                        "tax_rate": tax_rate,
                        "hs_code_id": hs,
                        "barcode": barcode,
                        "is_taxable": True,
                        "sale_price": Decimal(price),
                        "cost_price": Decimal(cost),
                    },
                )
                if was_created:
                    created += 1
                    if branch and not opts["no_stock"]:
                        record_movement(
                            tenant_id=tenant.id, product=Product.objects.get(tenant=tenant, sku=sku),
                            branch=branch, movement_type="opening_balance",
                            quantity=Decimal("50"))
                        stocked += 1
                else:
                    skipped += 1

        self.stdout.write(self.style.SUCCESS(
            f"Imported into {tenant.business_name}: {created} created, "
            f"{skipped} already existed, {stocked} stocked, "
            f"{unmapped} skipped (no valid HS code). "
            f"Barcodes {'kept' if opts['keep_barcodes'] else 'dropped (scan-to-assign)'}."))

    def _resolve_tenant(self, opts) -> Tenant:
        if opts.get("tenant_id"):
            try:
                return Tenant.objects.get(pk=opts["tenant_id"])
            except Tenant.DoesNotExist:
                raise CommandError("No tenant with that id.")
        if opts.get("ntn"):
            try:
                return Tenant.objects.get(ntn=opts["ntn"])
            except Tenant.DoesNotExist:
                raise CommandError("No tenant with that NTN.")
        raise CommandError("Pass --ntn or --tenant-id.")


def _slug(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", s.lower()).strip("-")[:60]
