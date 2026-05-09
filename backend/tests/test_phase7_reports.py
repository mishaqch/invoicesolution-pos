"""Phase 7 reports — registry, queries, exports, aggregates, isolation.

The strategy is to seed a tiny known fixture (a few invoices across two
days and one branch) and assert each report's numbers exactly. The
big-data perf assertion lives in test_phase7_perf for clarity.
"""

from __future__ import annotations

import datetime as dt
import io
import uuid
from decimal import Decimal

import pytest
from django.utils import timezone

from apps.catalog.models import Product, UnitOfMeasure
from apps.inventory.models import StockLevel, StockMovement
from apps.inventory.services.movements import record_movement
from apps.reports.aggregates import (
    rebuild_daily_sales,
    rebuild_product_velocity,
)
from apps.reports.base import BaseFilters
from apps.reports.exports import (
    excel_response,
    pdf_response,
    streaming_csv_response,
)
from apps.reports.models import DailySalesSummary, ProductVelocity, ReportFavorite
from apps.reports.registry import all_reports, get
from apps.sales.models import Invoice
from apps.sales.services import checkout
from apps.tenants.models import Branch, Tenant, Terminal


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def branch(db, tenant):
    return Branch.objects.create(
        tenant=tenant, name="Defence", code="RPT-1",
        address="x", city="x", province="SINDH",
    )


@pytest.fixture
def terminal(db, tenant, branch):
    return Terminal.objects.create(
        tenant=tenant, branch=branch, name="Counter A",
        device_fingerprint="rpt-fp",
    )


@pytest.fixture
def product(db, tenant, branch):
    p = Product.objects.create(
        tenant=tenant, name="Widget", sku="WGT-1",
        uom=UnitOfMeasure.objects.get(code="PCS"),
        sale_price=Decimal("100"),
        cost_price=Decimal("60"),
    )
    record_movement(
        tenant_id=tenant.id, product=p, branch=branch,
        movement_type="opening_balance", quantity=Decimal("100"),
    )
    return p


def _make_invoice(tenant, branch, terminal, cashier, product, *, qty=2, when=None):
    inv = checkout.create_invoice(
        tenant_id=tenant.id, branch=branch, terminal=terminal, cashier=cashier,
        cash_session=None, customer=None,
        cart_lines=[{
            "product": str(product.id),
            "quantity": str(qty),
            "unit_price": "100",
            "tax_rate": "18",
            "is_taxable": True,
        }],
        payments=[{"payment_method": "cash", "amount": str(Decimal(qty) * Decimal("118"))}],
        client_uuid=uuid.uuid4(),
    )
    # Aggregates only count statuses that "happened" — flip from
    # pending_sync to valid so reports see the row.
    fields = {"status": "valid"}
    if when:
        fields["invoice_date"] = when
    Invoice.objects.filter(pk=inv.pk).update(**fields)
    inv.refresh_from_db()
    return inv


# ---------------------------------------------------------------------------
# Registry
# ---------------------------------------------------------------------------


def test_all_18_reports_registered():
    expected = {
        "daily_sales", "hourly_heatmap", "item_wise", "category_wise",
        "top_movers", "slow_movers", "tax", "profit_loss", "stock",
        "stock_aging", "cashier_performance", "payment_breakdown",
        "returns_analysis", "customer_top_n", "customer_dormant",
        "supplier_purchase", "branch_comparison", "fbr_submissions",
        "audit_log",
    }
    # 19 actually because we ship 6 + 6 + 6 + audit_log + branch_comparison etc.
    # Verify the spec set of 18 is fully covered.
    assert expected.issubset(set(all_reports().keys()))


# ---------------------------------------------------------------------------
# Aggregates
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_rebuild_daily_sales_idempotent(tenant, branch, terminal, owner_user, product):
    today = timezone.localdate()
    _make_invoice(tenant, branch, terminal, owner_user, product, qty=2, when=today)
    _make_invoice(tenant, branch, terminal, owner_user, product, qty=1, when=today)

    rebuild_daily_sales(tenant, date_from=today, date_to=today)
    row1 = DailySalesSummary.objects.get(tenant=tenant, branch=branch, date=today)
    rebuild_daily_sales(tenant, date_from=today, date_to=today)
    row2 = DailySalesSummary.objects.get(tenant=tenant, branch=branch, date=today)

    # 2 invoices: 2*118 + 1*118 = 354 gross, 54 tax, 300 net.
    assert row1.invoice_count == 2
    assert row1.gross == Decimal("354.0000")
    assert row2.gross == row1.gross  # idempotent


@pytest.mark.django_db
def test_rebuild_product_velocity(tenant, branch, terminal, owner_user, product):
    today = timezone.localdate()
    _make_invoice(tenant, branch, terminal, owner_user, product, qty=3, when=today)
    rebuild_product_velocity(tenant, date_from=today, date_to=today)
    pv = ProductVelocity.objects.get(
        tenant=tenant, product=product, branch=branch, date=today,
    )
    assert pv.quantity == Decimal("3.0000")
    # revenue = 3 * 100 + 3 * 18 = 354 (line_total includes tax in this codebase)
    # cogs = 3 * 60 = 180
    assert pv.cogs == Decimal("180.0000")


# ---------------------------------------------------------------------------
# Reports — numbers match a hand-computed fixture
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_daily_sales_report_returns_correct_totals(
    tenant, branch, terminal, owner_user, product,
):
    today = timezone.localdate()
    _make_invoice(tenant, branch, terminal, owner_user, product, qty=2, when=today)
    _make_invoice(tenant, branch, terminal, owner_user, product, qty=1, when=today)
    rebuild_daily_sales(tenant, date_from=today, date_to=today)

    cls = get("daily_sales")
    rep = cls(tenant_id=str(tenant.id), filters=cls.Filters())
    result = rep.run(use_cache=False)
    assert result.row_count == 1
    assert result.rows[0]["invoice_count"] == 2
    assert result.totals["gross"] == Decimal("354.0000")


@pytest.mark.django_db
def test_item_wise_report(tenant, branch, terminal, owner_user, product):
    today = timezone.localdate()
    _make_invoice(tenant, branch, terminal, owner_user, product, qty=4, when=today)
    rebuild_product_velocity(tenant, date_from=today, date_to=today)

    cls = get("item_wise")
    rep = cls(tenant_id=str(tenant.id), filters=cls.Filters())
    rows = list(rep.run(use_cache=False).rows)
    assert len(rows) == 1
    assert rows[0]["sku"] == "WGT-1"
    assert rows[0]["quantity"] == Decimal("4.0000")


@pytest.mark.django_db
def test_payment_breakdown(tenant, branch, terminal, owner_user, product):
    today = timezone.localdate()
    _make_invoice(tenant, branch, terminal, owner_user, product, qty=2, when=today)

    cls = get("payment_breakdown")
    rep = cls(tenant_id=str(tenant.id), filters=cls.Filters(date_from=today, date_to=today))
    rows = list(rep.run(use_cache=False).rows)
    methods = {r["payment_method"]: r["total"] for r in rows}
    assert methods["cash"] == Decimal("236.0000")


@pytest.mark.django_db
def test_stock_report_filters_low(tenant, branch, terminal, owner_user, product):
    sl = StockLevel.objects.get(product=product, branch=branch)
    sl.reorder_level = Decimal("200")  # set a high reorder so we're "low"
    sl.save(update_fields=["reorder_level"])

    cls = get("stock")
    rep = cls(tenant_id=str(tenant.id), filters=cls.Filters(only_low=True))
    rows = list(rep.run(use_cache=False).rows)
    assert any(r["sku"] == "WGT-1" and r["status"] == "low" for r in rows)


@pytest.mark.django_db
def test_profit_loss_margin(tenant, branch, terminal, owner_user, product):
    today = timezone.localdate()
    _make_invoice(tenant, branch, terminal, owner_user, product, qty=5, when=today)
    rebuild_product_velocity(tenant, date_from=today, date_to=today)

    cls = get("profit_loss")
    rep = cls(tenant_id=str(tenant.id), filters=cls.Filters())
    result = rep.run(use_cache=False)
    # revenue = 5 * 118 = 590, cogs = 5 * 60 = 300, margin = 290
    assert result.totals["margin"] == Decimal("290.0000")


# ---------------------------------------------------------------------------
# Tenant isolation
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_reports_are_tenant_scoped(tenant, branch, terminal, owner_user, product):
    today = timezone.localdate()
    _make_invoice(tenant, branch, terminal, owner_user, product, qty=2, when=today)
    rebuild_daily_sales(tenant, date_from=today, date_to=today)

    other = Tenant.objects.create(
        business_name="Other", ntn=f"O-{uuid.uuid4().hex[:6]}",
        business_type="sole_proprietor", province="SINDH",
    )
    cls = get("daily_sales")
    rep = cls(tenant_id=str(other.id), filters=cls.Filters())
    result = rep.run(use_cache=False)
    assert result.row_count == 0


# ---------------------------------------------------------------------------
# Exports
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_csv_export_streams_money_formatted(
    tenant, branch, terminal, owner_user, product,
):
    today = timezone.localdate()
    _make_invoice(tenant, branch, terminal, owner_user, product, qty=2, when=today)
    rebuild_daily_sales(tenant, date_from=today, date_to=today)

    cls = get("daily_sales")
    result = cls(tenant_id=str(tenant.id), filters=cls.Filters()).run(use_cache=False)

    response = streaming_csv_response(result, filename="x.csv")
    body = b"".join(
        c if isinstance(c, bytes) else c.encode() for c in response.streaming_content
    ).decode()
    assert "Rs. 236.00" in body
    assert "Date" in body and "Branch" in body  # headers


@pytest.mark.django_db
def test_excel_export_round_trip(tenant, branch, terminal, owner_user, product):
    today = timezone.localdate()
    _make_invoice(tenant, branch, terminal, owner_user, product, qty=2, when=today)
    rebuild_daily_sales(tenant, date_from=today, date_to=today)

    cls = get("daily_sales")
    result = cls(tenant_id=str(tenant.id), filters=cls.Filters()).run(use_cache=False)
    response = excel_response(result, filename="x.xlsx")

    # Open the bytes back with openpyxl (read-only) and check shape.
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(response.content))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0][0] == "Date"  # header
    assert len(rows) >= 2  # header + data


@pytest.mark.django_db
def test_pdf_export_renders(tenant, branch, terminal, owner_user, product):
    today = timezone.localdate()
    _make_invoice(tenant, branch, terminal, owner_user, product, qty=2, when=today)
    rebuild_daily_sales(tenant, date_from=today, date_to=today)

    cls = get("daily_sales")
    result = cls(tenant_id=str(tenant.id), filters=cls.Filters()).run(use_cache=False)
    response = pdf_response(
        result, filename="x.pdf", title="Daily Sales",
        tenant_business_name="Khalil GS", tenant_ntn="1234567",
    )
    assert response.content[:4] == b"%PDF"
    assert len(response.content) > 1000  # has actual content


# ---------------------------------------------------------------------------
# Favorites
# ---------------------------------------------------------------------------


@pytest.mark.django_db
def test_report_favorite_unique_per_user_report_label(tenant, owner_user):
    ReportFavorite.objects.create(
        tenant=tenant, user=owner_user,
        report_name="daily_sales", label="Last 30 days",
        filters_json={"date_from": "2026-04-01"},
    )
    with pytest.raises(Exception):
        ReportFavorite.objects.create(
            tenant=tenant, user=owner_user,
            report_name="daily_sales", label="Last 30 days",
            filters_json={"date_from": "2026-04-15"},
        )
